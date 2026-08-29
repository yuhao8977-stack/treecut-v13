# -*- coding: utf-8 -*-
"""Phase 4 Stage 1.5 — V1.1 DELTA MERGE（V1.1 作为主源重导，含 P4 结构）。

输入：TreeCut_V11_Phase4.xlsx（知识总表 186 + P4_核心Taxonomy 92 + P4_映射与负向规则 30
      + P4_Evidence与Business 51 + 14_模板槽位 CT01-12 + 变更清单）
纪律：
  - 以 V1.1 的 knowledge_type 为准（用户已修正 BUSINESS_RULE 分类）
  - 不重复导入 V1.0 旧表（V1.1 已含全部 186）
  - CT06-CT12 导入为 HYPOTHESIS/DRAFT/UNVALIDATED
  - source_requirement_class 重分类（EXTERNAL/INTERNAL/PLATFORM/PRESENT/NO_NEED）
"""
import hashlib
import io
import json
import os
import sqlite3
import sys
from collections import Counter, defaultdict
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

REPO = r"C:\Users\admin\github\treecut-v13"
V11 = os.path.join(REPO, "knowledge", "source", "TreeCut_V11_Phase4.xlsx")
DATA_ROOT = r"E:\树剪整理\02_安装程序\TreeCut_v13\runtime_data\temp\batch1"

# V1.1 source_type → 归一
ST_MAP = {
    "业务词典": "business_dictionary",
    "运营业务模型": "operational_model",
    "运营假设": "operational_model",
    "生产规则": "production_rule",
    "视频生产模型": "production_rule",
    "内容生产模型": "production_rule",
    "平台官方": "platform_official",
    "专业机构参考": "professional_institution",
}

# source_requirement_class 分类规则（STEP 5-9）
def src_req_class(ktype, ns, source_type, statement):
    if ktype == "PLATFORM_RULE":
        return "PLATFORM_SOURCE_REQUIRED"
    if source_type == "professional_institution" or ns == "dimensions_decisions":
        return "EXTERNAL_SOURCE_REQUIRED"
    if ktype == "FACT" and "具体" in (statement or "") or any(k in (statement or "") for k in
            ("尺寸", "规范", "安全标准", "承重", "耐热", "耐刮", "环保", "电气")):
        return "EXTERNAL_SOURCE_REQUIRED"
    if source_type in ("operational_model", "production_rule"):
        return "INTERNAL_VALIDATION_REQUIRED"
    if source_type == "business_dictionary":
        return "SOURCE_PRESENT"
    return "NO_EXTERNAL_SOURCE_NEEDED"


def conf_tier(c):
    try:
        c = float(c)
    except Exception:
        return "UNKNOWN"
    if c >= 0.9:
        return "HIGH"
    if c >= 0.8:
        return "MEDIUM"
    if c >= 0.6:
        return "LOW"
    return "VERY_LOW"


def main():
    import openpyxl
    wb = openpyxl.load_workbook(V11, read_only=True, data_only=True)

    records = []

    # ---------- 1) 知识总表（186） ----------
    ws = wb["知识总表"]
    rows = list(ws.iter_rows(values_only=True))
    h = rows[0]
    idx = {str(x): i for i, x in enumerate(h)}
    for r in rows[1:]:
        if not r or not r[idx["knowledge_id"]]:
            continue
        kid = str(r[idx["knowledge_id"]]).strip()
        ktype = str(r[idx["knowledge_type"]] or "FACT").strip().upper()
        if ktype not in ("FACT", "BUSINESS_RULE", "HYPOTHESIS", "PLATFORM_RULE"):
            ktype = "FACT"
        ns = str(r[idx["namespace"]] or "").strip()
        st_s = str(r[idx["source_type"]] or "").strip()
        st = ST_MAP.get(st_s, st_s or "UNKNOWN")
        statement = str(r[idx["statement"]] or "").strip()
        src_cls = src_req_class(ktype, ns, st, statement)
        status = str(r[idx["status"]] or "DRAFT").strip().upper()
        # HYPOTHESIS 强制 DRAFT
        if ktype == "HYPOTHESIS" and status not in ("DRAFT",):
            status = "DRAFT"
        rec = {
            "knowledge_id": kid, "namespace": ns, "knowledge_type": ktype,
            "title": str(r[idx["name"]] or "").strip(),
            "statement": statement,
            "structured_payload": {
                "category": str(r[idx["category"]] or "").strip(),
                "aliases": [a.strip() for a in str(r[idx["aliases"]] or "").split("|") if a.strip()],
                "positive_evidence": str(r[idx["positive_evidence"]] or "").strip(),
                "exclude_evidence": str(r[idx["exclude_evidence"]] or "").strip(),
                "treecut_usage": str(r[idx["treecut_usage"]] or "").strip(),
                "content_roles": [x.strip() for x in str(r[idx["content_roles"]] or "").split("|") if x.strip()],
            },
            "source": str(r[idx["source_id"]] or st), "source_type": st,
            "source_version": str(r[idx["version"]] or "1.0").strip(),
            "confidence": conf_tier(r[idx["confidence"]]),
            "status": status,
            "effective_date": "2026-08-26",
            "expires_at": None,
            "ttl_days": int(str(r[idx["ttl"]]).replace("天", "").strip()) if r[idx["ttl"]] else None,
            "tags": [ns, ktype, str(r[idx["category"]] or "").strip()],
            "related_entities": [],
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "updated_at": None, "supersedes": None, "superseded_by": None,
            "source_requirement_class": src_cls,
            "validation_status": "UNVALIDATED" if ktype == "HYPOTHESIS" else
                                ("REVIEWED" if status == "REVIEWED" else "VALIDATED" if status == "ACTIVE" else "DRAFT"),
            "future_validation": "HISTORICAL_PERFORMANCE_TEST" if ktype == "HYPOTHESIS" else None,
            "needs_external_verification": src_cls == "EXTERNAL_SOURCE_REQUIRED",
            "section": str(r[idx["section"]] or "").strip(),
            "review_note": str(r[idx["review_note"]] or "").strip(),
        }
        records.append(rec)

    # ---------- 2) P4_核心Taxonomy（92） ----------
    ws2 = wb["P4_核心Taxonomy"]
    rows2 = list(ws2.iter_rows(values_only=True))
    for i, r in enumerate(rows2[1:], 1):
        if not r or not r[0]:
            continue
        group = str(r[0]).strip()
        code = str(r[1]).strip()
        name = str(r[2] or "").strip()
        ktype = str(r[4] or "BUSINESS_RULE").strip().upper()
        status = str(r[5] or "DRAFT").strip().upper()
        ns_map = {"USER_NEED": "user_needs", "BUSINESS_VALUE": "business_value_rules",
                  "CONTENT_ROLE": "content_roles", "MOTHER_THEME": "mother_themes",
                  "SHOT_FUNCTION": "shot_ontology", "SEARCH_INTENT": "semantic_mappings",
                  "DECISION_FACTOR": "dimensions_decisions", "CONTENT_TYPE": "content_types"}
        ns = ns_map.get(group, group.lower())
        if ktype not in ("FACT", "BUSINESS_RULE", "HYPOTHESIS", "PLATFORM_RULE"):
            ktype = "BUSINESS_RULE"
        if ktype == "HYPOTHESIS":
            status = "DRAFT"
        records.append({
            "knowledge_id": f"P4-{group}-{code}", "namespace": ns, "knowledge_type": ktype,
            "title": name or code, "statement": str(r[3] or "").strip(),
            "structured_payload": {"group": group, "code": code, "note": str(r[6] or "").strip()},
            "source": "USER_CURATED_STRUCTURED_KB", "source_type": "internal_business_model",
            "source_version": "1.1", "confidence": "MEDIUM", "status": status,
            "effective_date": "2026-08-29", "expires_at": None, "ttl_days": None,
            "tags": [group, ktype], "related_entities": [],
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "updated_at": None, "supersedes": None, "superseded_by": None,
            "source_requirement_class": "INTERNAL_VALIDATION_REQUIRED",
            "validation_status": "UNVALIDATED" if ktype == "HYPOTHESIS" else "DRAFT",
            "future_validation": "HISTORICAL_PERFORMANCE_TEST" if ktype == "HYPOTHESIS" else None,
            "needs_external_verification": False,
            "section": "P4_CORE_TAXONOMY", "review_note": "V1.1 Phase4 新增 Taxonomy"})

    # ---------- 3) P4_映射与负向规则（30） ----------
    ws3 = wb["P4_映射与负向规则"]
    rows3 = list(ws3.iter_rows(values_only=True))
    map_cnt = 0
    for r in rows3[2:]:
        if not r or not r[0]:
            continue
        mid = str(r[0]).strip()
        if not mid.startswith(("MAP", "NR")):
            continue
        map_cnt += 1
        ns = "semantic_mappings" if mid.startswith("MAP") else "negative_rules"
        ktype = "BUSINESS_RULE"
        records.append({
            "knowledge_id": mid, "namespace": ns, "knowledge_type": ktype,
            "title": mid, "statement": f"{r[1]} -> {r[2]}",
            "structured_payload": {"pattern": str(r[1] or ""), "meaning": str(r[2] or ""),
                                   "reliability": str(r[3] or ""), "note": str(r[4] or "")},
            "source": "USER_CURATED_STRUCTURED_KB", "source_type": "internal_business_model",
            "source_version": "1.1", "confidence": "MEDIUM",
            "status": "ACTIVE" if ns == "negative_rules" else "DRAFT",
            "effective_date": "2026-08-29", "expires_at": None, "ttl_days": None,
            "tags": [ns, ktype], "related_entities": [],
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "updated_at": None, "supersedes": None, "superseded_by": None,
            "source_requirement_class": "NO_EXTERNAL_SOURCE_NEEDED",
            "validation_status": "SYSTEM_GUARDRAIL" if ns == "negative_rules" else "DRAFT",
            "future_validation": None, "needs_external_verification": False,
            "section": "P4_MAPPING_NEGATIVE", "review_note": "V1.1 新增映射/负规则"})

    # ---------- 4) P4_Evidence与Business（51） ----------
    ws4 = wb["P4_Evidence与Business"]
    rows4 = list(ws4.iter_rows(values_only=True))
    ev_cnt = 0
    for i, r in enumerate(rows4[2:], 1):
        if not r or not r[0]:
            continue
        src = str(r[0]).strip()
        # 剔除表头误读行（如 title=confidence / evidence_source 是说明行）
        if src.lower() in ("confidence", "evidence_source", "reliability", "phase3依据/定位", "phase4使用限制"):
            continue
        rel = str(r[1] or "").strip()
        ev_cnt += 1
        records.append({
            "knowledge_id": f"ER-{i:03d}", "namespace": "semantic_mappings",
            "knowledge_type": "BUSINESS_RULE", "title": src, "statement": f"reliability={rel}",
            "structured_payload": {"evidence_source": src, "reliability": rel,
                                   "basis": str(r[2] or ""), "limit": str(r[3] or "")},
            "source": "USER_CURATED_STRUCTURED_KB", "source_type": "system_design",
            "source_version": "1.1", "confidence": "MEDIUM", "status": "ACTIVE",
            "effective_date": "2026-08-29", "expires_at": None, "ttl_days": None,
            "tags": ["EVIDENCE_RELIABILITY"], "related_entities": [],
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "updated_at": None, "supersedes": None, "superseded_by": None,
            "source_requirement_class": "NO_EXTERNAL_SOURCE_NEEDED",
            "validation_status": "SYSTEM_DESIGN", "future_validation": None,
            "needs_external_verification": False,
            "section": "P4_EVIDENCE_RELIABILITY", "review_note": "V1.1 Evidence Reliability Policy"})

    # ---------- 5) 模板（14_模板槽位 CT01-12） ----------
    ws5 = wb["14_模板槽位"]
    rows5 = list(ws5.iter_rows(values_only=True))
    h5 = rows5[0]
    i5 = {str(x): i for i, x in enumerate(h5)}
    templates = defaultdict(lambda: {"slots": [], "meta": {}})
    for r in rows5[1:]:
        if not r or not r[i5["template_id"]]:
            continue
        tid = str(r[i5["template_id"]]).strip()
        t = templates[tid]
        t["meta"] = {"template_id": tid,
                     "name": str(r[i5["template_name"]] or "").strip(),
                     "validation_status": str(r[i5["validation_status"]] or "").strip(),
                     "knowledge_type": str(r[i5["knowledge_type"]] or "HYPOTHESIS").strip().upper(),
                     "version": str(r[i5["version"]] or "1.0").strip(),
                     "source": str(r[i5["source"]] or "").strip()}
        t["slots"].append({"order": r[i5["顺序"]], "slot": str(r[i5["槽位"]] or "").strip(),
                           "beat": str(r[i5["脚本节拍"]] or "").strip(),
                           "target": str(r[i5["目标画面"]] or "").strip(),
                           "semantics": str(r[i5["必需语义"]] or "").strip(),
                           "preferred_exclude": str(r[i5["优选镜头 / 排除"]] or "").strip(),
                           "duration": str(r[i5["时长"]] or "").strip(),
                           "role": str(r[i5["内容角色"]] or "").strip()})
    for tid, t in templates.items():
        ktype = t["meta"]["knowledge_type"]
        if ktype not in ("FACT", "BUSINESS_RULE", "HYPOTHESIS", "PLATFORM_RULE"):
            ktype = "HYPOTHESIS"
        is_new = str(t["meta"]["source"]).startswith("V1.1") or tid >= "CT06"
        status = "DRAFT" if (ktype == "HYPOTHESIS" or is_new) else "ACTIVE"
        records.append({
            "knowledge_id": f"TPL-{tid}", "namespace": "template_library",
            "knowledge_type": ktype, "title": f"{tid} {t['meta']['name']}",
            "statement": f"模板 {tid}：{t['meta']['name']}，{len(t['slots'])} 槽位",
            "structured_payload": {"template_meta": t["meta"], "slots": t["slots"]},
            "source": str(t["meta"]["source"]), "source_type": "internal_business_model",
            "source_version": t["meta"]["version"], "confidence": "MEDIUM",
            "status": status, "effective_date": "2026-08-29", "expires_at": None,
            "ttl_days": None, "tags": ["TEMPLATE", tid], "related_entities": [],
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "updated_at": None, "supersedes": None, "superseded_by": None,
            "source_requirement_class": "NO_EXTERNAL_SOURCE_NEEDED",
            "validation_status": "UNVALIDATED" if ktype == "HYPOTHESIS" else
                                ("REVIEWED_SEED" if str(t["meta"]["validation_status"]) == "REVIEWED_SEED" else "DRAFT"),
            "future_validation": "HISTORICAL_PERFORMANCE_TEST" if ktype == "HYPOTHESIS" else None,
            "needs_external_verification": False,
            "section": "14_模板槽位", "review_note": f"CT01-05 原V1.0；CT06-12 V1.1 DRAFT"})

    wb.close()

    # ---------- 去重（knowledge_id 唯一）+ 统计 ----------
    seen = {}
    for r in records:
        seen[r["knowledge_id"]] = r
    records = list(seen.values())
    print("合并后总条数:", len(records))
    print("knowledge_type:", dict(Counter(r["knowledge_type"] for r in records)))
    print("namespace:", dict(Counter(r["namespace"] for r in records)))
    print("source_req:", dict(Counter(r["source_requirement_class"] for r in records)))
    print("status:", dict(Counter(r["status"] for r in records)))
    print("模板:", len(templates), "| MAP/NR:", map_cnt, "| EvidencePolicy:", ev_cnt)

    # ---------- 写 namespace 文件 + SQLite + manifest ----------
    by_ns = defaultdict(list)
    for r in records:
        by_ns[r["namespace"]].append(r)
    for ns, recs in by_ns.items():
        d = os.path.join(REPO, "knowledge", ns)
        os.makedirs(d, exist_ok=True)
        json.dump({"namespace": ns, "count": len(recs), "records": recs,
                   "source": "V1.1 Phase4 结构化补全版"},
                  open(os.path.join(d, "knowledge.json"), "w", encoding="utf-8"),
                  ensure_ascii=False, indent=1)

    db = os.path.join(DATA_ROOT, "knowledge_brain.db")
    conn = sqlite3.connect(db)
    conn.execute("DROP TABLE IF EXISTS knowledge_entries")
    conn.execute("""CREATE TABLE knowledge_entries(
        knowledge_id TEXT PRIMARY KEY, namespace TEXT, knowledge_type TEXT,
        title TEXT, statement TEXT, structured_payload TEXT,
        source TEXT, source_type TEXT, source_version TEXT,
        confidence TEXT, status TEXT, effective_date TEXT, expires_at TEXT,
        ttl_days INTEGER, tags TEXT, related_entities TEXT,
        created_at TEXT, updated_at TEXT, supersedes TEXT, superseded_by TEXT,
        needs_source INTEGER, note TEXT, source_sha256 TEXT,
        source_requirement_class TEXT, validation_status TEXT, future_validation TEXT,
        needs_external_verification INTEGER, section TEXT)""")
    conn.execute("DELETE FROM knowledge_entries")
    for r in records:
        conn.execute("""INSERT INTO knowledge_entries VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                     (r["knowledge_id"], r["namespace"], r["knowledge_type"], r["title"],
                      r["statement"], json.dumps(r["structured_payload"], ensure_ascii=False),
                      r["source"], r["source_type"], r["source_version"], r["confidence"],
                      r["status"], r["effective_date"], r["expires_at"], r["ttl_days"],
                      json.dumps(r["tags"]), json.dumps(r["related_entities"]),
                      r["created_at"], r["updated_at"], r["supersedes"], r["superseded_by"],
                      1 if r["needs_external_verification"] else 0, r["review_note"], None,
                      r["source_requirement_class"], r["validation_status"], r["future_validation"],
                      1 if r["needs_external_verification"] else 0, r["section"]))
    conn.commit()
    conn.close()
    print("SQLite 更新:", len(records))

    canon = json.dumps(records, ensure_ascii=False, sort_keys=True)
    manifest = {"manifest": "KNOWLEDGE_MANIFEST_V1_1",
                "source": {"file": "knowledge/source/TreeCut_V11_Phase4.xlsx",
                           "sha256": "07AE586D8655F5BB09EAD77012B5595D42AD09B03D00E3F4B0D302CBEABD7C0C",
                           "source_type": "USER_CURATED_STRUCTURED_KB",
                           "extends": "TreeCut_行业认知知识库_V1.0"},
                "record_count": len(records),
                "by_type": dict(Counter(r["knowledge_type"] for r in records)),
                "by_namespace": dict(Counter(r["namespace"] for r in records)),
                "by_source_req": dict(Counter(r["source_requirement_class"] for r in records)),
                "by_status": dict(Counter(r["status"] for r in records)),
                "template_count": len(templates),
                "records_sha256": hashlib.sha256(canon.encode("utf-8")).hexdigest(),
                "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M")}
    mp = os.path.join(REPO, "knowledge", "knowledge_manifest.json")
    json.dump(manifest, open(mp, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print("->", mp)
    print("records_sha256:", manifest["records_sha256"])


if __name__ == "__main__":
    main()
