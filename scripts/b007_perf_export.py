# -*- coding: utf-8 -*-
"""V0.2 — B007 Creator Performance 官方导出捕获（Source Route A 优先）。

流程：启动浏览器 → note-manager → 已发布 tab → 挂载 download 监听 → 点「下载/导出」按钮
→ 保存官方导出文件 → 解析 xlsx/csv → 记录实际列/行数/日期范围。
DOM 校准：语义文本（导出/下载报表/下载数据/导出报表/下载），禁固定坐标。
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from treecut.browser.config import load_config
from treecut.browser.main import BrowserRuntime

NOTE_MANAGER = "https://creator.xiaohongshu.com/new/note-manager"
TRIGGER_TEXTS = ("导出", "下载报表", "下载数据", "导出报表", "下载")


def main(argv=None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(line_buffering=True)
    ap = argparse.ArgumentParser()
    ap.add_argument("--workspace", default="B007")
    ap.add_argument("--headless", action="store_true")
    args = ap.parse_args(argv)

    config = load_config()
    config.workspace_id = args.workspace
    config.validate()
    runtime = BrowserRuntime(config)
    try:
        runtime.workspace.acquire_lock()
    except RuntimeError as error:
        print(f"PROFILE_LOCKED: {error}")
        return 2
    try:
        runtime.start_browser(headless=args.headless)
        result = runtime._in_browser(lambda: run_export(runtime), timeout=600)
        out = Path(r"C:\Users\admin\github\treecut-v13\reports\storage\B007_PERF_EXPORT_V1.json")
        out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    finally:
        runtime.close()
        print("EXPORT_DONE")


def run_export(runtime) -> dict:
    tab = runtime.ensure_tabs().get("CREATOR")
    downloads = []
    tab.on("download", lambda dl: downloads.append(dl))
    try:
        tab.goto(NOTE_MANAGER, timeout=60000)
        time.sleep(4)
        # 已发布 tab
        try:
            tab.evaluate(
                "() => { const els = Array.from(document.querySelectorAll('div,span,li,a,button'));"
                " const t = els.find(e => (e.textContent||'').trim() === '已发布' && e.children.length <= 2);"
                " if (t) { t.click(); return true; } return false; }")
            time.sleep(2)
        except Exception:
            pass
        tab.reload(timeout=60000)
        time.sleep(5)

        # 找导出/下载按钮（语义文本 + 稳定属性）
        found = tab.evaluate(
            """(texts) => {
              const els = Array.from(document.querySelectorAll('button,a,span,div,[class*=download],[class*=export],[class*=btn]'));
              for (const e of els) {
                const t = (e.textContent||'').trim();
                if (texts.some(x => t === x || t.startsWith(x)) && t.length <= 12) {
                  const r = e.getBoundingClientRect();
                  if (r.width > 0 && r.height > 0) {
                    e.click();
                    return {text: t, cls: (e.className||'').toString().slice(0,80)};
                  }
                }
              }
              return null;
            }""", TRIGGER_TEXTS)
        print("CLICKED =", found)
        if not found:
            return {"status": "EXPORT_BUTTON_NOT_FOUND",
                    "note": "未找到导出/下载按钮（语义文本扫描失败）"}

        # 等待下载（最多 90s）
        dl = None
        deadline = time.time() + 90
        while time.time() < deadline:
            if downloads:
                dl = downloads[0]
                break
            time.sleep(1.0)
        if dl is None:
            return {"status": "NO_DOWNLOAD_EVENT",
                    "clicked": found,
                    "note": "点击后 90s 内无 download 事件（可能弹窗/需要选择范围）"}
        # 落盘
        exports = Path(runtime.workspace.workspace_dir) / "treecut_inbox" / "creator" / "raw" / "creator" / "exports"
        exports.mkdir(parents=True, exist_ok=True)
        fname = dl.suggested_filename or f"creator_export_{int(time.time())}.xlsx"
        target = exports / fname
        dl.save_as(str(target))
        print("DOWNLOADED =", target)

        # 解析
        info = parse_export(target)
        return {"status": "EXPORT_DOWNLOADED", "file": str(target),
                "filename": fname, "size_bytes": target.stat().st_size,
                "parsed": info}
    except Exception as e:
        return {"status": "EXPORT_ERROR", "error": str(e)[:300]}


def parse_export(path: Path) -> dict:
    suffix = path.suffix.lower()
    info = {"format": suffix, "columns": [], "row_count": 0, "date_range": None,
            "sample_rows": []}
    try:
        if suffix == ".csv":
            import csv as _csv
            with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
                rows = list(_csv.reader(f))
            if rows:
                info["columns"] = rows[0]
                info["row_count"] = len(rows) - 1
                info["sample_rows"] = rows[1:4]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                info["columns"] = [str(c) if c is not None else "" for c in rows[0]]
                info["row_count"] = len(rows) - 1
                info["sample_rows"] = [[str(c)[:40] if c is not None else "" for c in r]
                                       for r in rows[1:4]]
        # 日期范围（找含时间的列）
        if info["sample_rows"]:
            cols = info["columns"]
            for i, cname in enumerate(cols):
                if any(k in cname.lower() for k in ("时间", "日期", "发布", "time", "date")):
                    vals = []
                    for r in info["sample_rows"]:
                        if i < len(r):
                            vals.append(r[i])
                    if vals:
                        info["date_range"] = {"col": cname, "sample": vals}
                    break
    except Exception as e:
        info["parse_error"] = str(e)[:200]
    return info


if __name__ == "__main__":
    sys.exit(main())
